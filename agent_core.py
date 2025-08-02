# agent_core.py
import re
import datetime as dt
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
import plotly.express as px
import plotly.graph_objects as go


# ------------------------------- Data model -------------------------------
@dataclass
class Activity:
    name: str
    area: str
    trade: str
    description: str
    start_day: int
    end_day: int
    duration: int
    # CPM (relative day index, 1-based)
    es: Optional[int] = None
    ef: Optional[int] = None
    ls: Optional[int] = None
    lf: Optional[int] = None
    slack: Optional[int] = None
    is_critical: bool = False
    # Calendar dates (set once start date known)
    start_date: Optional[dt.date] = None
    end_date: Optional[dt.date] = None


# ------------------------------- Excel helpers -------------------------------
def _find_header_row(ws) -> int:
    """Find the row that has 'Day 1' (case-insensitive). Fallback to a row with ≥2 'Day n' cells."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and re.match(r"(?i)^day\s*1$", v.strip()):
                return r
    for r in range(1, ws.max_row + 1):
        hits = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and re.match(r"(?i)^day\s*\d+$", v.strip()):
                hits += 1
        if hits >= 2:
            return r
    return 2


def _day_columns(ws, header_row: int) -> Tuple[List[int], List[str]]:
    day_cols, labels = [], []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and re.match(r"(?i)^day\s*\d+$", v.strip()):
            day_cols.append(c)
            labels.append(v.strip())
    return day_cols, labels


# ------------------------------- Main parser -------------------------------
def parse_excel_schedule(path: str) -> Tuple[List[Activity], Dict[str, int]]:
    """
    Auto-scan all sheets and pick the one with the most 'Day n' headers.
    Column A is the Area (merged cells handled by fill-down).
    Non-empty cells under Day columns form contiguous segments -> activities.
    Trade = prefix before ':' in cell text; description = full text of first non-empty cell in the segment.
    """
    wb = load_workbook(path, data_only=True)

    def score(ws):
        hr = _find_header_row(ws)
        dcols, labels = _day_columns(ws, hr)
        return len(dcols), hr, dcols, labels

    best_ws = None
    best_meta = None
    best_score = -1
    for ws in wb.worksheets:
        s, hr, dcols, labels = score(ws)
        if s > best_score:
            best_ws, best_meta, best_score = ws, (hr, dcols, labels), s

    if best_ws is None or best_score <= 0:
        return [], {"num_days": 0, "header_row": None, "day_labels": []}

    ws = best_ws
    header_row, day_cols, labels = best_meta
    num_days = len(day_cols)

    activities: List[Activity] = []
    current_area: Optional[str] = None

    for r in range(header_row + 1, ws.max_row + 1):
        raw_area = ws.cell(r, 1).value
        if raw_area not in (None, ""):
            current_area = str(raw_area).strip()
        if not current_area:
            continue  # skip until we hit a non-empty area cell

        # Collect text across day columns
        texts = []
        flags = []
        for c in day_cols:
            v = ws.cell(r, c).value
            txt = str(v).strip() if isinstance(v, str) else ""
            texts.append(txt)
            flags.append(bool(txt))

        # Contiguous segments of non-empty cells
        segs: List[Tuple[int, int]] = []
        in_seg = False
        s = None
        for j, flag in enumerate(flags, start=1):
            if flag and not in_seg:
                in_seg, s = True, j
            elif not flag and in_seg:
                segs.append((s, j - 1))
                in_seg = False
        if in_seg:
            segs.append((s, len(flags)))

        # Build activities from segments
        for (sd, ed) in segs:
            trade, desc = "General", ""
            for d in range(sd, ed + 1):
                cell_txt = texts[d - 1]
                if cell_txt:
                    desc = cell_txt
                    m = re.match(r"^\s*([A-Za-z ]+)\s*:", cell_txt)
                    if m:
                        trade = m.group(1).strip().title()
                    break
            name = f"{current_area} — {trade}"
            activities.append(
                Activity(
                    name=name,
                    area=current_area,
                    trade=trade,
                    description=desc,
                    start_day=sd,
                    end_day=ed,
                    duration=ed - sd + 1,
                )
            )

    return activities, {"num_days": num_days, "header_row": header_row, "day_labels": labels}


# ------------------------------- Dependencies & CPM -------------------------------
def infer_sequential_dependencies(n: int) -> Dict[int, List[int]]:
    """Default: purely sequential. Replace later with a real precedence editor."""
    deps = {i: [] for i in range(n)}
    for i in range(1, n):
        deps[i - 1].append(i)
    return deps


def compute_cpm(acts: List[Activity], deps: Dict[int, List[int]]) -> None:
    preds: Dict[int, List[int]] = {i: [] for i in range(len(acts))}
    for u, succs in deps.items():
        for v in succs:
            preds[v].append(u)

    # Forward
    for i, a in enumerate(acts):
        a.es = 1 if not preds[i] else max(acts[p].ef for p in preds[i]) + 1
        a.ef = a.es + a.duration - 1

    # Backward
    project_end = max((a.ef for a in acts), default=0)
    for i in reversed(range(len(acts))):
        a = acts[i]
        succs = deps.get(i, [])
        a.lf = project_end if not succs else min(acts[s].ls for s in succs)
        a.ls = a.lf - a.duration + 1
        a.slack = a.ls - a.es
        a.is_critical = (a.slack == 0)


def apply_calendar(acts: List[Activity], start_date: dt.date) -> None:
    for a in acts:
        a.start_date = start_date + dt.timedelta(days=a.es - 1)
        a.end_date = start_date + dt.timedelta(days=a.ef - 1)


# ------------------------------- Costs & EV -------------------------------
def estimate_hours_and_cost(
    acts: List[Activity],
    hours_per_day: float,
    base_rate_inr: float,
    labour_burden: float,
    inefficiency: float,
) -> Tuple[float, float]:
    total_hours = sum(a.duration for a in acts) * hours_per_day
    loaded_rate = base_rate_inr * (1 + labour_burden) * (1 + inefficiency)
    total_cost = total_hours * loaded_rate
    return total_hours, total_cost


def compute_contingency(cost: float, contingency: float) -> float:
    return cost * contingency


def earned_value(
    acts: List[Activity],
    today: dt.date,
    hours_per_day: float,
    base_rate_inr: float,
    labour_burden: float,
    inefficiency: float,
    pct_by_task: Dict[str, float],
) -> Dict[str, float]:
    loaded_rate = base_rate_inr * (1 + labour_burden) * (1 + inefficiency)
    PV = EV = AC = 0.0
    for a in acts:
        budget = a.duration * hours_per_day * loaded_rate
        if a.end_date < today:
            planned_pct = 1.0
        elif a.start_date > today:
            planned_pct = 0.0
        else:
            planned_days = (today - a.start_date).days + 1
            planned_pct = min(1.0, max(0.0, planned_days / a.duration))
        actual_pct = min(1.0, max(0.0, pct_by_task.get(a.name, 0.0) / 100.0))
        PV += budget * planned_pct
        EV += budget * actual_pct
        AC += budget * actual_pct  # proxy until actual hours are tracked
    SPI = (EV / PV) if PV > 0 else 0.0
    CPI = (EV / AC) if AC > 0 else 0.0
    return {"PV": PV, "EV": EV, "AC": AC, "SV": EV - PV, "CV": EV - AC, "SPI": SPI, "CPI": CPI}


# --------- Inefficiency analytics (for crew batching & gap reduction) ---------
def daily_workload_by_trade(acts: List[Activity], start_date: dt.date) -> pd.DataFrame:
    rows = []
    for a in acts:
        for d in range(a.es, a.ef + 1):
            rows.append({"date": start_date + dt.timedelta(days=d - 1), "trade": a.trade, "task": a.name})
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["date", "trade", "task_count"])
    return df.groupby(["date", "trade"]).agg(task_count=("task", "nunique")).reset_index()


def gaps_by_area(acts: List[Activity]) -> pd.DataFrame:
    df = pd.DataFrame([{"area": a.area, "es": a.es, "ef": a.ef, "name": a.name} for a in acts]).sort_values(["area", "es"])
    rows = []
    for area, g in df.groupby("area"):
        prev_ef = None
        for _, r in g.iterrows():
            if prev_ef is not None and r["es"] > prev_ef + 1:
                rows.append({"area": area, "gap_days": int(r["es"] - prev_ef - 1), "after_task": r["name"]})
            prev_ef = r["ef"]
    return pd.DataFrame(rows).sort_values(["area", "gap_days"], ascending=[True, False])


def fragmentation_by_trade(acts: List[Activity]) -> pd.DataFrame:
    df = pd.DataFrame([{"trade": a.trade, "name": a.name} for a in acts])
    if df.empty:
        return pd.DataFrame(columns=["trade", "segments"])
    return df.groupby("trade").agg(segments=("name", "nunique")).reset_index().sort_values("segments", ascending=False)


def improvement_suggestions(acts: List[Activity], start_date: dt.date) -> List[str]:
    tips = []
    wl = daily_workload_by_trade(acts, start_date)
    if not wl.empty:
        peaks = wl.groupby("trade")["task_count"].max().sort_values(ascending=False)
        tips.append("Trade load peaks (max concurrent tasks): " + ", ".join([f"{t}: {c}" for t, c in peaks.head(3).items()]))
    gaps = gaps_by_area(acts)
    if not gaps.empty:
        tips.append("Largest idle gaps by area: " + "; ".join([f"{r.area}: {r.gap_days}d" for r in gaps.head(5).itertuples()]))
    frag = fragmentation_by_trade(acts)
    if not frag.empty:
        tips.append("Most fragmented trades: " + ", ".join([f"{r.trade} ({int(r.segments)})" for r in frag.head(3).itertuples()]))
    crit = sum(1 for a in acts if a.is_critical)
    if crit:
        tips.append(f"{crit} tasks on critical path — prioritise these to compress duration.")
    return tips


# ------------------------------- Clean Gantt -------------------------------
def make_gantt(acts: List[Activity], pct_by_task: Dict[str, float], height_per_task: int = 24) -> go.Figure:
    rows = []
    for a in acts:
        rows.append({"Task": a.name, "Type": "Planned", "Start": a.start_date, "Finish": a.end_date,
                     "Area": a.area, "Trade": a.trade, "Critical": a.is_critical, "Percent": 100})
        pct = min(100.0, max(0.0, pct_by_task.get(a.name, 0.0)))
        days = round(pct / 100.0 * a.duration)
        prog_end = max(a.start_date, a.start_date + dt.timedelta(days=max(0, days - 1)))
        rows.append({"Task": a.name, "Type": "Progress", "Start": a.start_date, "Finish": prog_end,
                     "Area": a.area, "Trade": a.trade, "Critical": a.is_critical, "Percent": pct})
    df = pd.DataFrame(rows)
    if df.empty:
        return go.Figure()

    fig = px.timeline(
        df, x_start="Start", x_end="Finish", y="Task", color="Type",
        color_discrete_map={"Planned": "#D3D3D3", "Progress": "#2ca02c"},
        hover_data=["Area", "Trade", "Percent", "Start", "Finish"],
    )
    fig.update_yaxes(autorange="reversed")

    # % labels on progress bars
    for tr in fig.data:
        if tr.name == "Progress":
            mask = (df["Type"] == "Progress")
            tr.text = [f"{p:.0f}%" for p in df.loc[mask, "Percent"]]
            tr.textposition = "inside"
            tr.insidetextanchor = "middle"

    # outline planned bars (critical) in red
    for tr in fig.data:
        if tr.name == "Planned":
            tr.marker.line.color = "red"
            tr.marker.line.width = 1.2

    # today line
    today = dt.date.today()
    fig.add_shape(type="line", x0=today, x1=today, y0=-0.5, y1=len(df["Task"].unique()) + 0.5,
                  line=dict(color="orange", width=2, dash="dash"))

    fig.update_layout(
        height=max(420, min(1400, df["Task"].nunique() * height_per_task)),
        margin=dict(l=200, r=40, t=50, b=40),
        title="Gantt (Planned vs Progress) — critical outlined in red",
        legend_title_text="",
    )
    return fig
